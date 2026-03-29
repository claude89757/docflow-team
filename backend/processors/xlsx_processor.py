from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font


class XlsxProcessor:
    """Excel (.xlsx) 文档处理器

    只处理文本类型的 cell，保留公式、数字、日期等不动。
    """

    def parse(self, file_path: str) -> dict:
        """解析 xlsx 文档结构，只返回文本 cell"""
        wb = load_workbook(file_path, data_only=False)
        sheets = []
        for ws in wb.worksheets:
            text_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "s" and cell.value:
                        text_cells.append(
                            {
                                "row": cell.row,
                                "col": cell.column,
                                "col_letter": cell.column_letter,
                                "value": cell.value,
                                "font_name": cell.font.name if cell.font else None,
                                "font_size": cell.font.size if cell.font else None,
                                "bold": cell.font.bold if cell.font else None,
                            }
                        )
            sheets.append(
                {
                    "name": ws.title,
                    "row_count": ws.max_row or 0,
                    "col_count": ws.max_column or 0,
                    "text_cells": text_cells,
                }
            )
        wb.close()
        return {
            "file_path": file_path,
            "format": "xlsx",
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
        }

    def extract_text(self, file_path: str) -> str:
        """提取所有文本 cell 内容"""
        wb = load_workbook(file_path, data_only=True)
        texts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "s" and cell.value and cell.value.strip():
                        texts.append(cell.value)
        wb.close()
        return "\n".join(texts)

    def get_stats(self, file_path: str) -> dict:
        """文档统计"""
        wb = load_workbook(file_path, data_only=False)
        text_cell_count = 0
        total_chars = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "s" and cell.value:
                        text_cell_count += 1
                        total_chars += len(cell.value)
        result = {
            "sheet_count": len(wb.sheetnames),
            "text_cell_count": text_cell_count,
            "total_chars": total_chars,
        }
        wb.close()
        return result

    def replace_text(self, file_path: str, replacements: dict, output_path: str) -> str:
        """替换文本 cell 内容，保留格式

        replacements 格式: {"sheet_name:row:col": "new_text"}
        或 {"sheet_idx:row:col": "new_text"} (sheet_idx 为数字)
        遇到公式 cell 自动跳过。
        """
        wb = load_workbook(file_path)

        for key, new_text in replacements.items():
            parts = str(key).split(":")
            if len(parts) != 3:
                continue
            sheet_ref, row_str, col_str = parts
            row, col = int(row_str), int(col_str)

            # 查找 worksheet
            ws = self._find_sheet(wb, sheet_ref)
            if ws is None:
                continue

            cell = ws.cell(row=row, column=col)
            # 跳过公式 cell
            if cell.data_type == "f":
                continue
            # 只替换文本 cell
            if cell.data_type == "s" or cell.value is None:
                old_font = copy(cell.font) if cell.font else None
                cell.value = new_text
                if old_font:
                    cell.font = old_font

        wb.save(output_path)
        wb.close()
        return output_path

    def apply_format_changes(self, file_path: str, changes: list[dict], output_path: str) -> str:
        """应用格式修改指令到文本 cell

        changes 格式:
        [
            {"target": "font_size", "scope": "all", "value": 12},
            {"target": "font_name", "scope": "Sheet1", "value": "微软雅黑"},
            {"target": "bold", "scope": "all", "value": true},
        ]
        scope: "all" 或 sheet 名称
        """
        wb = load_workbook(file_path)

        for change in changes:
            target = change.get("target")
            scope = change.get("scope", "all")

            sheets = wb.worksheets if scope == "all" else [self._find_sheet(wb, scope)]
            sheets = [ws for ws in sheets if ws is not None]

            for ws in sheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type != "s" or not cell.value:
                            continue

                        if target == "font_size":
                            cell.font = Font(
                                name=cell.font.name,
                                size=change["value"],
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                color=cell.font.color,
                            )
                        elif target == "font_name":
                            cell.font = Font(
                                name=change["value"],
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                color=cell.font.color,
                            )
                        elif target == "bold":
                            cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=change["value"],
                                italic=cell.font.italic,
                                color=cell.font.color,
                            )

        wb.save(output_path)
        wb.close()
        return output_path

    def _find_sheet(self, wb, ref: str):
        """按名称或索引查找 worksheet"""
        if ref in wb.sheetnames:
            return wb[ref]
        try:
            idx = int(ref)
            if 0 <= idx < len(wb.worksheets):
                return wb.worksheets[idx]
        except ValueError:
            pass
        return None
