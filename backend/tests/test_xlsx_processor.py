"""XlsxProcessor 单元测试"""

import os
import tempfile

from openpyxl import Workbook, load_workbook

from backend.processors.xlsx_processor import XlsxProcessor


def _create_simple_xlsx(path: str, data: dict[str, list[list]]):
    """创建测试用 xlsx

    data: {"Sheet1": [["标题", "值"], ["描述", 42], ...]}
    """
    wb = Workbook()
    first = True
    for sheet_name, rows in data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


def test_parse():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test.xlsx")
        _create_simple_xlsx(path, {"Sheet1": [["标题", "描述"], ["数据", 100]]})

        result = processor.parse(path)
        assert result["format"] == "xlsx"
        assert result["sheet_count"] == 1
        assert len(result["sheets"][0]["text_cells"]) >= 3  # 标题, 描述, 数据


def test_extract_text():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test.xlsx")
        _create_simple_xlsx(path, {"Sheet1": [["测试文本", "更多文本"], ["第二行", 42]]})

        text = processor.extract_text(path)
        assert "测试文本" in text
        assert "更多文本" in text
        assert "第二行" in text
        assert "42" not in text  # 数字不应出现


def test_get_stats():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test.xlsx")
        _create_simple_xlsx(path, {"Sheet1": [["文本A", "文本B"], ["文本C", 100]]})

        stats = processor.get_stats(path)
        assert stats["sheet_count"] == 1
        assert stats["text_cell_count"] == 3
        assert stats["total_chars"] > 0


def test_replace_text():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        src = os.path.join(tmpdir, "src.xlsx")
        out = os.path.join(tmpdir, "out.xlsx")
        _create_simple_xlsx(src, {"Sheet1": [["原始标题", "原始描述"], ["保持不变", 42]]})

        # 替换 Sheet1 的 row=1,col=1 和 row=1,col=2
        processor.replace_text(src, {"Sheet1:1:1": "新标题", "Sheet1:1:2": "新描述"}, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "新标题"
        assert ws.cell(row=1, column=2).value == "新描述"
        assert ws.cell(row=2, column=1).value == "保持不变"
        assert ws.cell(row=2, column=2).value == 42  # 数字不变
        wb.close()


def test_replace_text_skips_formula():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        src = os.path.join(tmpdir, "src.xlsx")
        out = os.path.join(tmpdir, "out.xlsx")

        # 手动创建带公式的 xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "文本"
        ws["B1"] = 10
        ws["C1"] = "=B1*2"  # 公式
        wb.save(src)
        wb.close()

        # 尝试替换公式 cell，应该被跳过
        processor.replace_text(src, {"Sheet1:1:3": "不应替换"}, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws["C1"].value == "=B1*2"  # 公式保持不变
        wb.close()


def test_apply_format_changes():
    processor = XlsxProcessor()
    with tempfile.TemporaryDirectory() as tmpdir:
        src = os.path.join(tmpdir, "src.xlsx")
        out = os.path.join(tmpdir, "out.xlsx")
        _create_simple_xlsx(src, {"Sheet1": [["文本A", "文本B"]]})

        changes = [{"target": "font_size", "scope": "all", "value": 14}]
        processor.apply_format_changes(src, changes, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.size == 14
        wb.close()
